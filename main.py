import time
import asyncio
import sqlite3
from aiogram import Bot, Dispatcher, types, executor
import openai
import re
client = openai.OpenAI(api_key="sk-proj-dMxBPSgPFHxHooKvFEMmT3BlbkFJDEuEtpmEZMkxJbHrwvvk")
from aiogram.types import ChatActions
users = []
from openpyxl import Workbook, load_workbook

file1 = client.files.create(
  file=open("ТРЕБОВАНИЯ К СПИКЕРАМ.pdf", "rb"),
  purpose='assistants'
)
file2 = client.files.create(
  file=open("Основные_правила.txt", "rb"),
  purpose='assistants'
)
file3 = client.files.create(
  file=open("Ценности_сообщества_Код_публичности.docx", "rb"),
  purpose='assistants'
)


Assistant_ID = 'asst_mYsdSqR87HqRc6hkmVIuU01x'
print(Assistant_ID)
TELEGRAM_TOKEN = '7114081148:AAHlZfYWMxEng3mQZeP9zyqdfLLCvy-aJLw'

bot = Bot(token=TELEGRAM_TOKEN)
dp = Dispatcher(bot)
threads = {}

async def remove_text_in_brackets(text):
    # Используем регулярное выражение для поиска текста в квадратных скобках и удаляем его
    return re.sub(r'\[.*?\]', '', text)
async def write_to_excel(file_name, values):
    try:
        # Пытаемся загрузить существующую книгу Excel
        wb = load_workbook(file_name)
    except FileNotFoundError:
        # Если файл не найден, создаем новую книгу
        wb = Workbook()

    # Выбираем активный лист или создаем новый, если его нет
    ws = wb.active

    # Записываем значения в новую строку
    ws.append(values)

    # Сохраняем книгу
    wb.save(file_name)
values_to_write = ['дата', 'ник', 'вопрос', 'ответ']  # Значения для записи
write_to_excel("output.xlsx", values_to_write)

@dp.message_handler(commands=['send'])
async def send_excel_file(message: types.Message):
    file_name = "output.xlsx"
    with open(file_name, 'rb') as file:
        await bot.send_document(message.chat.id, file)
# Функция для добавления пользователя в базу данных
async def handle_with_assistant(message, chat_id):
    await bot.send_chat_action(message.chat.id, ChatActions.TYPING)
    print('генерация началась')

    thread_id = threads[chat_id]
    message_answer = client.beta.threads.messages.create(
        thread_id=thread_id,
        role="user",
        content=f"{message.text}"

    )
    run = client.beta.threads.runs.create(
        thread_id=thread_id,
        assistant_id=Assistant_ID,

    )

    time.sleep(10)
    run_status = client.beta.threads.runs.retrieve(
        thread_id=thread_id,
        run_id=run.id
    )

    print(run_status.status)
    while run_status.status == 'in_progress':
        time.sleep(5)
        run_status = client.beta.threads.runs.retrieve(
            thread_id=thread_id,
            run_id=run.id
        )
    print(run_status.status)



    if run_status.status == 'completed':
        messages = client.beta.threads.messages.list(
            thread_id=thread_id
        )

        msg = messages.data[0]
        role = msg.role
        content = msg.content[0].text.value
        print(f"{role.capitalize()}: {content}")

        fixed_content = await remove_text_in_brackets(content)
        fixed_content = re.sub(r"\\\(.*?\\\)", lambda m: eval(m.group(0)[2:-2]), fixed_content)

        # Убираем LaTeX-спецсимволы
        fixed_content = re.sub(r"\\,|\\text|{|}", "", fixed_content)

        # Убираем лишние символы и пробелы
        fixed_content =  fixed_content.replace("(", "").replace(")", "").replace("\\", "").strip()
        await bot.send_message(chat_id=message.chat.id, text=fixed_content)
        date = message.date.strftime("%Y-%m-%d %H:%M:%S")
        username = message.from_user.username
        question = message.text
        answ = content
        values_to_write.clear()
        values_to_write.append(date)
        values_to_write.append(username)
        values_to_write.append(question)
        values_to_write.append(answ)
        await write_to_excel("output.xlsx", values_to_write)
        values_to_write.clear()


def add_user(chat_id):

    thread = client.beta.threads.create()
    threads[chat_id] = thread.id
    print(thread.id)
    return thread.id


async def answer_user(message_response, message):
    await message.answer(message_response)


@dp.message_handler(commands=['start'])
async def start_command(message: types.Message):
    add_user(message.chat.id)
    await message.answer("Привет! Вы зарегистрированы.")


@dp.message_handler(content_types=types.ContentTypes.TEXT)
async def echo_message(message: types.Message):
    if message.text != '/send':
        await handle_with_assistant(message, message.chat.id)


# Запуск бота
if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
